#==========================================================================
# Program: FindCircuitsEquipmentFacilities.py ---- CLASS ---
# Author:  Jorge E. Rodriguez
# Date Created: Feb-17-2018
# Date Last Modified: Feb-18-2018
# Summary: This is Class to for the Circuits
#==========================================================================

#***************************************************************
# ==================== Libraries Required <BEGIN> =============*
#***************************************************************

#************************ For PING ************************
import re  # Required for teh Class
import subprocess
from time import time, sleep
try:
    import socket
    import threading
#    fromthreading import *
except:
    print ("NO Sockets is available")
#************************ For PING ************************

import os
import sys
import math
import datetime
import time
import random
import tkinter
import tkinter.messagebox
import tkinter.filedialog
from tkinter import *           # Importing the Tkinter (tool box) library
from tkinter import ttk
if sys.version_info < (3,0): 
    import Tkinter as tkinter 
    import tkMessageBox as mbox 
    import Tkinter.font as tkfont
else: 
    import tkinter 
    import tkinter.messagebox as mbox 
    import tkinter.font as tkfont
#import PyPDF2

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
    Is_Excell_Available = True        
except:
    print ("********************************************************************************* \n")
    print ("*** No openpyxl library exist, please make sure you downlaod it and instal it *** \n")
    print ("********************************************************************************* \n")
    Is_Excell_Available = False    

try:
    from odbc_connector import *
    Is_ODBC_Available = True
except:
    print ("********************************************************************************** \n")
    print ("*** NO ODBC Library Found, please download it in order to access the Databases *** \n")
    print ("********************************************************************************** \n")
    Is_ODBC_Available = False
    sys.exit(1)

#*******************************************************
#============= READING VARIABLES TABLE ================*
#*******************************************************
try:
    from Utils import *
    Utils = Class_Utils()
    Utils.Get_Values()
    #------- DNS NAME ---------
    ODBC_DSN_name = Utils.Get_ODBC_Name()
    Windows_Scaling = Utils.Get_Windows_Scaling()
    #--------------------------
except:
    #------- DNS NAME ---------
    ODBC_DSN_name = "BV"
    Windows_Scaling = 1.0
    #--------------------------

#print (Windows_Scaling)

try:
    from Logging import *
    Is_logging_Available = True
    Parameter = []
    Parameter = ['Report Outsource Cost Per Country','OPEN Window']    
    Logging = Class_Logging(ODBC_DSN_name,Parameter)
    Logging.Log(Parameter)
except:
    print ("********************************************************************************** \n")
    print ("*** NO Logging Library Found, please download it in order to access the Databases *** \n")
    print ("********************************************************************************** \n")
    Is_logging_Available = False

# ------------------ Colors -------------------
white = 'FFFFFF'
black = '000000'
indianred = 'CD5C5C'
indianred_1 = 'FF6A6A'
indianred_2 = 'EE6363'
indianred_4 = '8B3A3A'
indianred_3 = 'CD5555'
brown = 'A52A2A'
brown_1 = 'FF4040'
brown_2 = 'EE3B3B'
brown_3 = 'CD3333'
brown_4 = '8B2323'
firebrick = 'B22222'
firebrick_1 = 'FF3030'
firebrick_2 = 'EE2C2C'
firebrick_3 = 'CD2626'
firebrick_4 = '8B1A1A'
red_1 = 'FF0000'
red_2 = 'EE0000'
red_3 = 'CD0000'
chocolate = 'D2691E'
chocolate_1 = 'FF7F24'
chocolate_2 = 'EE7621'
chocolate_3 = 'CD661D'
chocolate_4 = '8B4513'
sienna = 'A0522D'
sienna_1 = 'FF8247'
sienna_2 = 'EE7942'
sienna_3 = 'CD6839'
sienna_4 = '8B4726'
lightsalmon_1 = 'FFA07A'
lightsalmon_2 = 'EE9572'
lightsalmon_3 = 'CD8162'
lightsalmon_4 = '8B5742'
orangered_1 = 'FF4500'
orangered_2 = 'EE4000'
orangered_3 = 'CD3700'
orangered_4 = '8B2500'
darkorange = 'FF8C00'
darkorange_1 = 'FF7F00'
darkorange_2 = 'EE7600'
darkorange_3 = 'CD6600'
darkorange_4 = '8B4500'
orange = 'FF8000'
gold_1 = 'FFD700'
gold_2 = 'EEC900'
gold_3 = 'CDAD00'
gold_4 = '8B7500'
yellow_1 = 'FFFF00'
yellow_2 = 'EEEE00'
yellow_3 = 'CDCD00'
yellow_4 = '8B8B00'
green = '008000'
green_1 = '00FF00'
green_2 = '00EE00'
green_3 = '00CD00'
green_4 = '008B00'
darkgreen = '006400'
sapgreen = '308014'
lawngreen = '7CFC00'
chartreuse_1 = '7FFF00'
chartreuse_2 = '76EE00'
chartreuse_3 = '66CD00'
chartreuse_4 = '458B00'
greenyellow = 'ADFF2F'
springgreen = '00FF7F'
springgreen_1 = '00EE76'
springgreen_2 = '00CD66'
springgreen_3 = '008B45'
mediumseagreen = '3CB371'
seagreen_1 = '54FF9F'
seagreen_2 = '4EEE94'
seagreen_3 = '43CD80'
seagreen_4 = '2E8B57'
cyan = '00FFFF'
cyan_2 = '00EEEE'
cyan_3 = '00CDCD'
cyan_4 = '008B8B'
dodgerblue_1 = '1E90FF'
dodgerblue_2 = '1C86EE'
dodgerblue_3 = '1874CD'
dodgerblue_4 = '104E8B'
aliceblue = 'F0F8FF'
steelblue = '4682B4'
steelblue_1 = '63B8FF'
steelblue_2 = '5CACEE'
steelblue_3 = '4F94CD'
steelblue_4 = '36648B'
lightskyblue = '87CEFA'
lightskyblue_1 = 'B0E2FF'
lightskyblue_2 = 'A4D3EE'
lightskyblue_3 = '8DB6CD'
lightskyblue_4 = '607B8B'
skyblue_1 = '87CEFF'
skyblue_2 = '7EC0EE'
skyblue_3 = '6CA6CD'
skyblue_4 = '4A708B'
skyblue = '87CEEB'
deepskyblue_1 = '00BFFF'
deepskyblue_2 = '00B2EE'
deepskyblue_3 = '009ACD'
deepskyblue_4 = '00688B'
blue = '0000FF'
blue_2 = '0000EE'
blue_3 = '0000CD'
blue_4  = '00008B'
navy = '000080'
indigo = '4B0082'
blueviolet = '8A2BE2'
purple_1 = '9B30FF'
purple_2 = '912CEE'
purple_3 = '7D26CD'
purple_4 = '551A8B'
violetred_1 = 'FF3E96'
violetred_2 = 'EE3A8C'
violetred_3 = 'CD3278'
violetred_4 = '8B2252'
turquoise_1 = '00F5FF'
turquoise_2 = '00E5EE'
turquoise_3 = '00C5CD'
turquoise_4 = '00868B'


#*********************************************************************************************************************************************
#                                   ReportOutsourcePerCountry Section <BEGIN>                                                                                   *
#*********************************************************************************************************************************************
class Class_SaveAs:

    def __init__(self,DSN_Name,Windows_Scaling,FileName,ToolVersion):
        self.ODBC_name = DSN_Name
        self.db = ODBC(self.ODBC_name)
        self.db2 = ODBC(self.ODBC_name)
        self.SaveAsWindowExist = False
        self.Username = os.getlogin()
        self.date = ""
        self.Windows_Scaling = Windows_Scaling
        self.data_ready = False
        self.FileName = FileName 
        #self.ColumnsName = ColumnsName
        #self.ColumnsData = ColumnsData
        self.ToolVersion = ToolVersion
        # --------- Excell ------------
        self.workbook = ""
        self.worksheet = ""

    #-------------------- Excell Section BEGIN -------------------

    def Build_Template_Tabs(self):
        worksheet = self.workbook.active
        # Change the name of the Sheet
        ss_sheet = self.workbook.get_sheet_by_name('Sheet')
        ss_sheet.title = self.TabsNames[0]
        # Create/Add a NEW Sheet
        i = 1
        while (i < len(self.TabsNames)):                
            self.workbook.create_sheet(self.TabsNames[i])
            i = i + 1

    def Tab_Style(self,worksheet,cell_range,Color,Size,Merge,Bold):        
        first_cell = worksheet[cell_range.split(":")[0]]
        if Merge == 'Merge':
            worksheet.merge_cells(cell_range)
            CellAlignment = Alignment(horizontal="center", vertical="center")
            first_cell.alignment = CellAlignment

        if Merge == 'Mergeleft':
            worksheet.merge_cells(cell_range)
            CellAlignment = Alignment(horizontal="left", vertical="center")
            first_cell.alignment = CellAlignment
                   
        fill = PatternFill("solid", fgColor=Color )
        if Bold == 'Bold':
            fontobj = Font(name='Calibri',size=Size,bold=True,color=white)
        else:
            fontobj = Font(name='Calibri',size=Size,bold=False,color=white)

        first_cell.font = fontobj
        first_cell.fill = fill


    def Build_Columns_Template(self):
        i = 0
        while (i < len(self.TabsNames)):
            SheetReport = self.workbook.get_sheet_by_name(self.TabsNames[i])
            i = i + 1
            #------ Repoprt 0 TAB ---------    
            SheetReport['A1'] = 'BVAnalytics:'
            self.Tab_Style(SheetReport,'A1:A1',steelblue,12,'NO','Bold')
            SheetReport['B1'] = self.ToolVersion
            self.Tab_Style(SheetReport,'B1:B1',steelblue,12,'NO','Bold')
            SheetReport['C1'] = 'Prepared By:'
            self.Tab_Style(SheetReport,'C1:C1',steelblue,12,'NO','Bold')
            SheetReport['D1'] = os.getlogin()
            self.Tab_Style(SheetReport,'D1:D1',steelblue,12,'NO','Bold')
            SheetReport['E1'] = 'Date and Time:'
            self.Tab_Style(SheetReport,'E1:E1',steelblue,12,'NO','Bold')
            SheetReport['F1'] = datetime.datetime.now()
            self.Tab_Style(SheetReport,'F1:F1',steelblue,12,'Merge','NO')
            self.Tab_Style(SheetReport,'G1:AI1',steelblue,12,'Merge','NO')
        
    def Add_DataToWorksheet(self,String,Row,Column,worksheet,Fill,Color,Size,Bold):

        fill = PatternFill("solid", fgColor=Color )
        if Bold == 'Bold':
            fontobj = Font(name='Calibri',size=Size,bold=True)
        else:
            fontobj = Font(name='Calibri',size=Size,bold=False)
        
        SheetReport = self.workbook.get_sheet_by_name(worksheet)
        j = 0
        while (j < len(String)):
            #SheetReport.cell(row=Row, column=Column, value=String[j])
            if (Fill == "Fill"):
                #SheetReport.cell(row=Row, column=Column, value=String[j])
                SheetReport.cell(row=Row, column=Column, value=String[j]).font = fontobj
                SheetReport.cell(row=Row, column=Column, value=String[j]).fill = fill
            else:
                SheetReport.cell(row=Row, column=Column, value=String[j]).font = fontobj
            j = j + 1
            Column = Column + 1

    def Add_DataToWorksheetSingleCell(self,String,Row,Column,worksheet,Fill,Color,Size,Bold):

        fill = PatternFill("solid", fgColor=Color )
        if Bold == 'Bold':
            fontobj = Font(name='Calibri',size=Size,bold=True)
        else:
            fontobj = Font(name='Calibri',size=Size,bold=False)
        
        SheetReport = self.workbook.get_sheet_by_name(worksheet)
        if (Fill == "Fill"):
            SheetReport.cell(row=Row, column=Column, value=String).font = fontobj
            SheetReport.cell(row=Row, column=Column, value=String).fill = fill
        else:
            SheetReport.cell(row=Row, column=Column, value=String).font = fontobj


    def Read_DataFromWorksheet(self,worksheet):
        Temporary_File_Name = self.FileName.replace(".","-BVAnalytics.")
        file_name = Temporary_File_Name
        Data = []
        try:
            self.workbook = load_workbook(self.FileName) # Open the Original File
            self.workbook.save(Temporary_File_Name)      # Save the Temp File so we can work it and not mess up teh Original
            self.FileName = Temporary_File_Name          # We change the self.FileName to be the Working File so we do not mess the Original ever
            self.workbook = load_workbook(self.FileName) # WE open the Temp File so all Routines use the Temp File.
            SheetReport = self.workbook.get_sheet_by_name(worksheet)
            #print (SheetReport)
            Row = 1
            Column = 1
            EndOfRows = True
            EndOfColumns = True
            #============================ Find the Total No. of Columns (consecutive Columns) =====================
            while EndOfColumns: 
                if (SheetReport.cell(row=Row, column=Column).value != None):
                    Total_No_Of_Columns = Column
                    Column = Column + 1
                else:
                    EndOfColumns = False
            Column = 1
            #============================ Find the Total No. of Rows (consecutive Rows) ===========================
            while EndOfRows:            
                if (SheetReport.cell(row=Row, column=Column).value != None):
                    Total_No_Of_Rows = Row
                    Row = Row + 1
                else:
                    EndOfRows = False
            #=========================== Read The total Number of Rows and Columns ================================
            Row = 1
            Column = 1
            #print (Total_No_Of_Rows)
            #print (Total_No_Of_Columns)
            while (Row <= Total_No_Of_Rows):
                Column = 1
                RowData = []
                while (Column <= Total_No_Of_Columns): 
                    #print ("["+str(Row)+"]:["+str(Column)+"]"+str(SheetReport.cell(row=Row, column=Column).value))
                    if (SheetReport.cell(row=Row, column=Column).value != None):
                        RowData.append(SheetReport.cell(row=Row, column=Column).value)
                    else:
                        RowData.append("")
                    Column = Column + 1
                #print (RowData)
                Data.append(RowData)
                Column = 1
                Row = Row + 1             
            return Data
        except: 
            return Data


    def Save_File(self):
        try:
            self.workbook.save(self.FileName)
            return True
        except:
            return False        


    #-------------------- Excell Section END ---------------------        

    def Call_Write_to_File(self,TabsNames):        
        #print ("Write to file....")
        #print (os.path.basename(file_name))
        Username = os.getlogin()
        DateAndTime = datetime.datetime.now()
        self.TabsNames = TabsNames
        self.workbook = Workbook()
        self.Build_Template_Tabs()
        self.Build_Columns_Template()

#*********************************************************************************************************************************************
#                                   ReportOutsourcePerCountry Section <END>                                                                                   *
#*********************************************************************************************************************************************


def Main():
    print ("Testing the Save As Class....:")
    location = []
    #location = ['UNKNOWN','UNKNOWN','UNKNOWN','UNKNOWN']
    Names = ["Column 1","Column 2","Column 3","Column 4"]
    Data = ["Data 1","Data 2","Data 3","Data 4"]
    Tabs = ["Report 0","Report 1","Report 2","Report 3","Report 4"]
    File = "c:\\temp\\Network_Gear_Inventory_20180103.xlsx"
    #File = "c:\\temp\\jorge.xlsx"
    ExcellFile = Class_SaveAs("BV",Windows_Scaling,File,"Ver 4.0")

    ##############################################################################################
    ############# Write information to Exell File and Rename Tabs <BEGIN> ########################
    ##############################################################################################
    '''
    ExcellFile.Call_Write_to_File(Tabs)
    Row = 2    # 3
    Column = 1 # A
    ExcellFile.Add_DataToWorksheet(Names,Row,Column,"Report 0","Fill",lightsalmon_1,14,'Bold')
    Row = 3
    Column = 1
    ExcellFile.Add_DataToWorksheet(Data,Row,Column,"Report 0","NO",steelblue,12,'NO')
    Row = 4
    Column = 1
    ExcellFile.Add_DataToWorksheet(Data,Row,Column,"Report 0","NO",steelblue,12,'NO')
    ExcellFile.Save_File()
    '''
    ############################################################################################
    ############# Write information to Exell File and Rename Tabs <END> ########################
    ############################################################################################

    ###########################################################################################
    ############# Read information to Exell File and Read Tabs <BEGIN> ########################
    ###########################################################################################
    ExcellFileData = ExcellFile.Read_DataFromWorksheet("CMDB") # <- from the CMDB Tab
    if (len(ExcellFileData) > 0):
        print ("Readig the File Succesfully")
        No_of_Rows = len(ExcellFileData)
        No_of_Columns = len(ExcellFileData[0])
        print ("The Nummber of Columns is: [ " + str(No_of_Columns) + " ]")
        print ("The Nummber of Rows is:    [ " + str(No_of_Rows) + " ]")
        Column_Names = ExcellFileData[0]
        print (Column_Names)
        column = 0
        row = 0
        #print (ExcellFileData[1829][13])
        ExcellFile.Add_DataToWorksheetSingleCell(ExcellFileData[1829][13],1829,13,"CMDB","Fill",lightsalmon_1,14,'Bold') #<---- To Mark Cells
        #row = No_of_Rows
        while (row < No_of_Rows):
            column = 0
            while (column <No_of_Columns):                
                #print (ExcellFileData[row][column])
                column = column + 1
            row = row + 1
            print ("["+str(row)+"]:["+str(column)+"]")
        ExcellFile.Save_File()
        #ExcellFile.Read_DataFromWorksheet("Report 1")
        #ExcellFile.Read_DataFromWorksheet("Report 2")
    else:
        print ("error Opening the file or reading it")
    ###########################################################################################
    ############# Read information to Exell File and Read Tabs <BEGIN> ########################
    ###########################################################################################

if __name__ == '__main__':
    Main()
